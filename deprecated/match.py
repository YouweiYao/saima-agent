#!/usr/bin/env python3
"""
赛马智能体 v6 - 重构版
- 三大块流程：功能匹配 → 风险评估 → Excel输出
- 支持参数配置
- 支持静默模式
- 状态实时更新
"""
import argparse
import json
import openpyxl
import requests
import jieba
import re
import math
import time
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

# ==================== 配置 ====================
QWEN_API_KEY = "sk-3d96ade0c8fa40378a4560fdd43b067e"
QWEN_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
DEFAULT_EXCEL_PATH = "/home/yaoyouwei/.openclaw/media/inbound/千帆融合版3.1功能清单---9a90c454-ccbf-48e1-957c-0878d2943a52.xlsx"
DEFAULT_REQ_PATH = "/home/openclaw/niuniu/saima/output/需求拆分_result.json"
DEFAULT_OUT_PATH = "/home/openclaw/niuniu/saima/output/final_output.json"
DEFAULT_OUT_EXCEL = "/home/openclaw/niuniu/saima/output/赛马匹配结果.xlsx"
DEFAULT_STATUS_FILE = "/tmp/saima_status.json"

# 产品映射
SHEETS = {"千帆appbuilder功能清单": "千帆AB", "千帆modelbuilder功能清单": "千帆MB"}

# 过滤需求类型
EXCLUDE_CATEGORIES = ["商务需求", "运维/维保需求", "验收需求"]

# ==================== 导入Prompt ====================
from prompts.match_prompt import MATCH_SYSTEM_PROMPT, build_match_user_prompt
from prompts.risk_prompt import RISK_SYSTEM_PROMPT, build_risk_user_prompt


# ==================== 工具函数 ====================
def update_status(status_file, stage, progress=None, total=None, detail=""):
    """更新状态文件"""
    try:
        data = {"status": "running", "stage": stage, "time": time.strftime("%Y-%m-%d %H:%M:%S")}
        if progress is not None:
            data["progress"] = progress
        if total is not None:
            data["total"] = total
        if detail:
            data["detail"] = detail
        with open(status_file, "w") as f:
            json.dump(data, f, ensure_ascii=False)
    except:
        pass

def log(verbose, *args, **kwargs):
    """日志输出（根据verbose控制）"""
    if verbose:
        print(*args, **kwargs)

def tokenize(text):
    return [w for w in jieba.cut(re.sub(r'[^\w\u4e00-\u9fff]', ' ', text)) if len(w) >= 2]

def get_caps(ws):
    caps = []
    for row in range(2, ws.max_row + 1):
        path = []
        for col in range(2, 7):
            v = ws.cell(row, col).value
            if v and str(v).strip() and str(v) != "-":
                path.append(str(v).strip())
        path_str = " > ".join(path)
        desc = ws.cell(row, 7).value or ""
        if path_str and desc:
            caps.append({"path": path_str, "desc": desc})
    return caps

def recall(text, caps, top_k=10):
    tokens = tokenize(text)
    scores = []
    for cap in caps:
        s = sum(1 for t in tokens if t in cap['path'].lower() or t in cap['desc'].lower())
        scores.append(s)
    indexed = list(enumerate(scores))
    indexed.sort(key=lambda x: -x[1])
    return [caps[i] for i, _ in indexed[:top_k]]

def call_llm(system_prompt, user_prompt, retries=3):
    """调用LLM API"""
    for attempt in range(retries):
        try:
            resp = requests.post(
                QWEN_BASE_URL + "/chat/completions",
                headers={"Authorization": f"Bearer {QWEN_API_KEY}"},
                json={"model": "qwen-plus", "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ], "temperature": 0.1},
                timeout=60
            )
            if resp.status_code == 200:
                content = resp.json()["choices"][0]["message"]["content"]
                # 提取JSON
                start = content.find('{')
                end = content.rfind('}') + 1
                if start >= 0 and end > start:
                    return json.loads(content[start:end])
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1)
    return None


# ==================== 步骤1: 功能匹配 ====================
def match_single_product(args):
    """单个产品功能匹配"""
    req_text, category, source, product, top_k, threshold, verbose = args
    
    caps = product["caps"]
    recalled = recall(req_text, caps, top_k)
    
    # 使用Prompt模板
    user_prompt = build_match_user_prompt(
        {"requirement": req_text, "category": category, "source_text": source},
        recalled,
        threshold
    )
    
    mf = call_llm(MATCH_SYSTEM_PROMPT, user_prompt)
    if not mf:
        return None
    
    score = float(mf.get("founction_match_level", "0").replace("%", "").strip() or "0")
    
    log(verbose, f"  {product['short']}: {'是' if mf.get('is_product_function_matched') == '是' else '否'} (得分: {score})")
    
    return {"mf": mf, "score": score, "caps": recalled, "product": product["short"]}

def step1_match(requirements, products, args):
    """步骤1: 功能匹配"""
    update_status(args.status, "步骤1: 功能匹配", progress=0, total=len(requirements))
    
    # 准备匹配任务
    match_tasks = []
    for req in requirements:
        for p in products:
            match_tasks.append((
                req["requirement"],
                req["category"],
                req.get("source_text", ""),
                p,
                args.top_k,
                args.threshold,
                args.verbose
            ))
    
    log(args.verbose, f"总匹配任务数: {len(match_tasks)}")
    
    # 并发执行
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        match_results = list(executor.map(match_single_product, match_tasks))
    
    t1 = time.time()
    log(args.verbose, f"匹配耗时: {t1-t0:.2f}秒")
    
    # 格式兼容处理：支持 {"results": [...]} 或 [...]
    if isinstance(requirements, dict):
        if "results" in requirements:
            requirements = requirements["results"]
    
    # 标准化需求格式（兼容新旧格式）
    normalized = []
    for req in requirements:
        # 新格式：item本身就有requirement字段
        if "requirement" in req:
            normalized.append(req)
        # 旧格式：requirements字段
        elif "requirements" in req:
            for r in req["requirements"]:
                r["source_text"] = req.get("source_text", "")
                normalized.append(r)
    requirements = normalized
    
    # 整理结果
    results = []
    task_idx = 0
    for req in requirements:
        best_match = None
        best_score = -1
        best_product = None
        best_caps = None
        
        for p in products:
            result = match_results[task_idx]
            task_idx += 1
            if not result:
                continue
            if result["score"] > best_score:
                best_score = result["score"]
                best_match = result["mf"]
                best_product = result["product"]
                best_caps = result["caps"]
        
        results.append({
            "requirement": req["requirement"],
            "category": req["category"],
            "source_text": req.get("source_text", ""),
            "matched_functions": []
        })
        
        if best_match:
            if best_score >= args.threshold:
                best_match["is_product_function_matched"] = "是"
                best_match["delivery_type"] = "标品功能"
                results[-1]["matched_functions"].append(best_match)
                log(args.verbose, f"需求: {req['requirement'][:30]}... -> 已满足({best_product})")
            else:
                results[-1]["matched_functions"].append({
                    "product_name": best_product,
                    "product_function_level": "",
                    "product_detail_source_text": "",
                    "founction_match_level": str(best_score),
                    "is_product_function_matched": "否"
                })
                log(args.verbose, f"需求: {req['requirement'][:30]}... -> 不满足，待风险评估")
        
        # 更新进度
        update_status(args.status, "步骤1: 功能匹配", 
                     progress=len(results), total=len(requirements))
    
    return results


# ==================== 步骤2: 风险评估 ====================
def risk_single_req(args):
    """单个需求风险评估"""
    req_text, category, matched_result, verbose = args
    
    # 使用Prompt模板
    user_prompt = build_risk_user_prompt(
        {"requirement": req_text, "category": category},
        matched_result
    )
    
    risk_result = call_llm(RISK_SYSTEM_PROMPT, user_prompt)
    log(verbose, f"  风险评估完成")
    return risk_result

def step2_risk(results, args):
    """步骤2: 风险评估"""
    # 找出未匹配的需求
    risk_tasks = []
    for req_result in results:
        mf = req_result.get("matched_functions", [])
        if mf and mf[0].get("is_product_function_matched") == "否":
            risk_tasks.append((
                req_result["requirement"],
                req_result["category"],
                mf[0],
                args.verbose
            ))
    
    if not risk_tasks:
        log(args.verbose, "所有需求已匹配，无需风险评估")
        return results
    
    update_status(args.status, "步骤2: 风险评估", progress=0, total=len(risk_tasks))
    log(args.verbose, f"风险评估任务数: {len(risk_tasks)}")
    
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        risk_results = list(executor.map(risk_single_req, risk_tasks))
    
    t1 = time.time()
    log(args.verbose, f"风险评估耗时: {t1-t0:.2f}秒")
    
    # 更新结果
    task_idx = 0
    for req_result in results:
        mf = req_result.get("matched_functions", [])
        if mf and mf[0].get("is_product_function_matched") == "否":
            if task_idx < len(risk_results) and risk_results[task_idx]:
                mf[0].update(risk_results[task_idx])
            task_idx += 1
        
        update_status(args.status, "步骤2: 风险评估",
                     progress=task_idx, total=len(risk_tasks))
    
    return results


# ==================== 步骤3: Excel输出 ====================
def step3_output(results, args):
    """步骤3: Excel输出"""
    update_status(args.status, "步骤3: 生成Excel")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "匹配结果"
    
    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 表头
    headers = ["来源原文", "需求描述", "需求分类", "产品名称", "产品功能层级", 
               "功能原文", "功能匹配度", "是否匹配", "交付方式", "需求质量度",
               "定制化工作说明", "是否开放需求", "风险应对策略"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # 数据
    row = 2
    for req_result in results:
        source_text = req_result.get("source_text", "")
        requirement = req_result.get("requirement", "")
        category = req_result.get("category", "")
        
        for mf in req_result.get("matched_functions", []):
            ws.cell(row, 1, source_text).border = thin_border
            ws.cell(row, 2, requirement).border = thin_border
            ws.cell(row, 3, category).border = thin_border
            ws.cell(row, 4, mf.get("product_name", "")).border = thin_border
            ws.cell(row, 5, mf.get("product_function_level", "")).border = thin_border
            ws.cell(row, 6, mf.get("product_detail_source_text", "")).border = thin_border
            ws.cell(row, 7, mf.get("founction_match_level", "")).border = thin_border
            ws.cell(row, 8, mf.get("is_product_function_matched", "")).border = thin_border
            ws.cell(row, 9, mf.get("delivery_type", "")).border = thin_border
            ws.cell(row, 10, mf.get("reqirement_quality_level", "")).border = thin_border
            ws.cell(row, 11, mf.get("customized_work_details", "")).border = thin_border
            ws.cell(row, 12, mf.get("is_open_requirement", "")).border = thin_border
            ws.cell(row, 13, mf.get("risk_management_strategy", "")).border = thin_border
            
            # 设置行高
            ws.row_dimensions[row].height = 30
            row += 1
    
    # 列宽
    for col in range(1, 14):
        ws.column_dimensions[chr(64+col)].width = 20
    
    # ==================== 合并单元格逻辑（从左到右依次判断）====================
    # 业务规则：只有当前面所有列的值都相同时才合并当前列
    # 参考: business-logic.md "Step 6: 输出Excel"
    for col in range(1, 14):  # 13列
        merge_start = None
        prev_values = []
        
        for r in range(2, row):  # 从第2行开始（数据从第2行开始）
            if col == 1:
                # 第1列（来源原文）：只需要比较当前列
                current_value = ws.cell(r, col).value or ""
                prev_value = ws.cell(r-1, col).value or "" if r > 2 else ""
                
                if merge_start is None:
                    merge_start = r
                    prev_value = current_value
                elif current_value == prev_value:
                    continue  # 相同，继续合并
                else:
                    # 不相同，结束之前的合并
                    if r - merge_start > 1:  # 至少2行才合并
                        ws.merge_cells(start_row=merge_start, start_column=col, end_row=r-1, end_column=col)
                    merge_start = r
                    prev_value = current_value
            else:
                # 第2-13列：需要比较当前列和前面所有列
                current_row_values = [ws.cell(r, c).value or "" for c in range(1, col+1)]
                prev_row_values = [ws.cell(r-1, c).value or "" for c in range(1, col+1)] if r > 2 else []
                
                if merge_start is None:
                    merge_start = r
                    prev_values = current_row_values
                elif current_row_values == prev_values:
                    continue  # 前面所有列都相同，继续合并
                else:
                    # 不相同，结束之前的合并
                    if r - merge_start > 1:
                        ws.merge_cells(start_row=merge_start, start_column=col, end_row=r-1, end_column=col)
                    merge_start = r
                    prev_values = current_row_values
        
        # 处理最后一段合并
        if merge_start and row - merge_start > 1:
            ws.merge_cells(start_row=merge_start, start_column=col, end_row=row-1, end_column=col)
    
    # ==================== 合并单元格逻辑结束 ====================
    
    wb.save(args.output)
    log(args.verbose, f"Excel已保存: {args.output}")
    
    # 保存JSON
    with open(args.output_json, "w") as f:
        json.dump({"results": results}, f, ensure_ascii=False, indent=2)
    log(args.verbose, f"JSON已保存: {args.output_json}")
    
    return args.output


# ==================== 主函数 ====================
def main():
    # 启动时清理旧的状态文件和中间结果
    import os
    status_file = DEFAULT_STATUS_FILE
    if os.path.exists(status_file):
        os.remove(status_file)
        print(f"已清理: {status_file}")
    
    parser = argparse.ArgumentParser(description="赛马智能体 v6")
    parser.add_argument("--input", "-i", default=DEFAULT_REQ_PATH, help="需求文件路径")
    parser.add_argument("--output", "-o", default=DEFAULT_OUT_EXCEL, help="输出Excel路径")
    parser.add_argument("--output-json", default=DEFAULT_OUT_PATH, help="输出JSON路径")
    parser.add_argument("--excel", "-e", default=DEFAULT_EXCEL_PATH, help="功能清单Excel")
    parser.add_argument("--workers", "-w", type=int, default=15, help="并发数")
    parser.add_argument("--threshold", "-t", type=float, default=0.8, help="匹配阈值")
    parser.add_argument("--top-k", "-k", type=int, default=10, help="BM25召回数")
    parser.add_argument("--status", "-s", default=DEFAULT_STATUS_FILE, help="状态文件")
    parser.add_argument("--quiet", "-q", action="store_true", help="静默模式")
    parser.add_argument("--verbose", "-v", action="store_true", help="详细输出")
    
    args = parser.parse_args()
    
    # 状态文件初始化
    update_status(args.status, "初始化")
    
    log(not args.quiet, "=" * 50)
    log(not args.quiet, "赛马智能体 v6 (重构版)")
    log(not args.quiet, f"并发数: {args.workers}, 阈值: {args.threshold}")
    log(not args.quiet, "=" * 50)
    
    # 读取功能清单
    log(not args.quiet, "加载功能清单...")
    wb = openpyxl.load_workbook(args.excel)
    products = []
    for sheet_name in SHEETS:
        if sheet_name in wb.sheetnames:
            caps = get_caps(wb[sheet_name])
            products.append({"name": sheet_name, "short": SHEETS[sheet_name], "caps": caps})
            log(not args.quiet, f"产品: {SHEETS[sheet_name]}, 功能: {len(caps)}")
    
    # 读取需求
    log(not args.quiet, "加载需求...")
    with open(args.input, "r") as f:
        raw = json.load(f)
        if isinstance(raw, dict) and "results" in raw:
            data = raw["results"]
        else:
            data = raw
    
    # 过滤
    filtered = [item for item in data if item.get("category", "") not in EXCLUDE_CATEGORIES]
    log(not args.quiet, f"需求数: {len(data)} -> 过滤后: {len(filtered)}")
    
    # 执行三步骤
    t_total = time.time()
    
    results = step1_match(filtered, products, args)
    results = step2_risk(results, args)
    step3_output(results, args)
    
    t_end = time.time()
    
    # 完成
    update_status(args.status, "完成", detail=f"总耗时: {t_end-t_total:.2f}秒")
    log(not args.quiet, f"\n总耗时: {t_end-t_total:.2f}秒")
    log(not args.quiet, "完成!")


if __name__ == "__main__":
    main()
