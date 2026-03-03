#!/usr/bin/env python3
MAX_RETRIES = 3
"""赛马智能体 v5 - 并发版"""

import sys, json, openpyxl, requests, jieba, re, math
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

EXCLUDE_CATEGORIES = ["商务需求", "运维/维保需求", "验收需求"]
# 配置
QWEN_API_KEY = "sk-3d96ade0c8fa40378a4560fdd43b067e"
QWEN_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
EXCEL_PATH = "/home/yaoyouwei/.openclaw/media/inbound/千帆融合版3.1功能清单---9a90c454-ccbf-48e1-957c-0878d2943a52.xlsx"
REQ_PATH = "/home/openclaw/niuniu/saima/output/需求拆分_result.json"
OUT_PATH = "/home/openclaw/niuniu/saima/output/final_output.json"
OUT_EXCEL_PATH = "/home/openclaw/niuniu/saima/output/赛马匹配结果.xlsx"
TOP_K = 10
THRESHOLD = float(sys.argv[2]) if len(sys.argv) > 2 else 0.8
MAX_WORKERS = int(sys.argv[1]) if len(sys.argv) > 2 else 15

# 产品映射
SHEETS = {"千帆appbuilder功能清单": "千帆AB", "千帆modelbuilder功能清单": "千帆MB"}

def tokenize(text):
    return [w for w in jieba.cut(re.sub(r'[^\w\u4e00-\u9fff]', ' ', text)) if len(w) >= 2]

def get_caps(ws):
    caps = []
    for row in range(2, ws.max_row + 1):
        path = []
        for col in range(2, 7):
            v = ws.cell(row, col).value
            if v and str(v).strip() and str(v) != "-": 
                path.append(str(v).strip())
        path_str = " > ".join(path)
        desc = ws.cell(row, 7).value or ""
        if path_str and desc:
            caps.append({"path": path_str, "desc": desc})
    return caps

def recall(text, caps):
    tokens = tokenize(text)
    scores = []
    for cap in caps:
        s = sum(1 for t in tokens if t in cap['path'].lower() or t in cap['desc'].lower())
        scores.append(s)
    indexed = list(enumerate(scores))
    indexed.sort(key=lambda x: -x[1])
    return [caps[i] for i, _ in indexed[:TOP_K]]

def call_llm(prompt):
    try:
        resp = requests.post(f"{QWEN_BASE_URL}/chat/completions",
            headers={"Authorization": f"Bearer {QWEN_API_KEY}", "Content-Type": "application/json"},
            json={"model": "qwen-plus", "messages": [{"role": "user", "content": prompt}], "temperature": 0.1}, timeout=60)
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"]
    except: pass
    return None

def build_match_prompt(req, features):
    features_str = "\n\n".join(f"【候选功能 {i+1}】\n{f['path']}\n{f['desc']}" for i, f in enumerate(features))
    
    system_prompt = """你是一名专业的软件产品功能匹配分析助手。

你的职责是：基于给定的【用户需求】与【候选产品功能列表】，判断这些功能是否可以满足该需求，并给出结构化、可审计的判断结果。

你必须遵守以下原则：
- 只能基于输入的候选产品功能进行判断
- 不得编造、假设或引入任何未在候选功能中出现的能力
- 判断应保持克制、保守和工程视角
- 若证据不足，应明确给出无法匹配的结论

你是一个系统组件，不是聊天助手。"""

    user_prompt = f"""请根据以下输入信息，完成一次【需求与产品功能匹配判断】

【一、输入数据】
- source_text：整体需求背景
- requirement：当前评估的用户需求
- category：需求类型
- features：候选产品功能全集

--------------------------------
【source_text】
{req.get('source_text', '')}

【requirement】
{req['requirement']}

【category】
{req['category']}

【features（候选功能全集）】
{features_str}

【二、匹配规则】

1. 聚合匹配原则
- 一个requirement允许被多个候选功能点共同覆盖

2. founction_match_level 定义
- 表示：候选功能集合对需求语义要点的整体覆盖程度
- 取值范围：[0.00, 1.00]
- 类型：字符串

3. is_product_function_matched 判定规则
- founction_match_level ≥ {THRESHOLD} → 是
- 否则 → 否

--------------------------------
【三、输出格式（严格JSON）】

{{
  "requirement_id": "REQ-001",
  "requirement": "需求原文",
  "category": "类别",
  "matched_functions": [
    {{
      "product_name": "千帆AB/千帆MB",
      "product_function_level": "层级路径",
      "product_detail_source_text": "功能原文",
      "founction_match_level": "0.00-1.00",
      "is_product_function_matched": "是或否"
    }}
  ]
}}

⚠️ 必须一次性输出完整JSON"""

    return system_prompt, user_prompt

def build_risk_prompt(req, matched_result):
    is_matched = matched_result.get('is_product_function_matched', '否')
    detail = matched_result.get('product_detail_source_text', '')
    
    system_prompt = """你是一名具有丰富软件交付与实施经验的技术方案专家，长期负责金融、政务、工业、能源等行业的软件项目需求澄清、实施评估、交付方案设计与成本测算。

⚠️ 重要前提：
- 客户所属行业为：金融 / 政务 / 工业 / 能源
- 必须默认按照最复杂、最严格、最保守的应用场景进行评估
- 不得基于理想实施环境进行乐观估算

⚠️ 标品交付直通约束：
- 仅当 is_product_function_matched = "是" 时，才允许判定为标品交付"""

    user_prompt = f"""【requirement】
{req['requirement']}

【category】
{req['category']}

【is_product_function_matched】
{is_matched}

【product_detail_source_text】
{detail}

--------------------------------
【处理规则】

1️⃣ 若 is_product_function_matched = "是"
- delivery_type = "产品交付"
- is_open_requirement = "否"
- requirement_clarity_score = 0
- 其他字段为空

2️⃣ 仅当is_product_function_matched为"否"时，才允许执行以下评估

一、需求清晰度与风险判断
- 输出 requirement_clarity_score（0~1）
- clarity_risk_type 必须从以下枚举中选择：信息缺失 / 歧义过多 / 范围不清 / 行业隐含依赖 / 产品强依赖 / 无明显风险

二、定制化工作与成本评估（强制保守）

Step 1：复杂度分级
- 强制规则：涉及模型管理 → 中复杂度；涉及权限/合规 → 高复杂度

Step 2：工作内容拆分
- 需求澄清、设计、开发、集成、测试

Step 3：人天估算（X-Y区间）
- 低：开发≥3人天，测试≥1人天
- 中：开发≥8人天，测试≥3人天
- 高：开发≥15人天，测试≥5人天

customized_work_details 结构：
1️⃣ Summary（是否定制化、复杂度）
2️⃣ Work Breakdown（工作内容）
3️⃣ Man-day Estimation（开发/测试人天X-Y区间）

--------------------------------
【输出格式】
{{
  "matched_functions": [
    {{
      "delivery_type": "类型",
      "reqirement_quality_level": "0-1之间的小数",
      "customized_work_details": "工作内容",
      "is_open_requirement": "是或否",
      "risk_management_strategy": "风险策略",
      "requirement_clarity_score": "0-1之间的小数",
      "clarity_risk_type": "风险类型"
    }}
  ]
}}

⚠️ 必须输出JSON"""

    return system_prompt, user_prompt

# 并发调用匹配
def match_single_product(args):
    """并发处理单个产品匹配"""
    req_text, category, source, p = args
    caps = recall(req_text, p['caps'])
    sp, up = build_match_prompt({"requirement": req_text, "category": category, "source_text": source}, caps)
    resp = call_llm(sp + "\n\n" + up)
    
    if not resp:
        return None
    
    try:
        result = json.loads(resp[resp.find('{'):resp.rfind('}')+1])
        mfs = result.get('matched_functions', [])
        if mfs and isinstance(mfs, list):
            mf = mfs[0]
            mf['product_name'] = p['short']
            score_str = mf.get('founction_match_level', '0')
            try:
                score = float(score_str) if score_str else 0
            except:
                score = 0
            return {"mf": mf, "score": score, "caps": caps, "product": p['short']}
    except:
        pass
    return None

# 并发调用风险评估
def risk_single_req(args):
    """并发处理单个需求的风险评估"""
    req_text, category, empty_match = args
    rsp, rup = build_risk_prompt({"requirement": req_text, "category": category}, empty_match)
    rresp = call_llm(rsp + "\n\n" + rup)
    
    if rresp:
        try:
            r = json.loads(rresp[rresp.find('{'):rresp.rfind('}')+1])
            if r.get('matched_functions'):
                return r['matched_functions'][0]
        except:
            pass
    return None

def main():
    print("=" * 50)
    print("赛马智能体 v5 (并发版)")
    print(f"并发数: {MAX_WORKERS}")
    print("=" * 50)
    
    # 读取产品
    wb = openpyxl.load_workbook(EXCEL_PATH)
    products = []
    for sheet_name in SHEETS:
        if sheet_name in wb.sheetnames:
            caps = get_caps(wb[sheet_name])
            products.append({"name": sheet_name, "short": SHEETS[sheet_name], "caps": caps})
            print(f"产品: {SHEETS[sheet_name]}, 功能: {len(caps)}")
    
    # 读取需求
    with open(REQ_PATH, 'r') as f:
        raw = json.load(f)
        # 兼容两种格式：{"results": [...]} 或 [...]
        if isinstance(raw, dict) and 'results' in raw:
            data = raw['results']
        else:
            data = raw
    
    # 过滤掉不需要的需求类型（根据业务逻辑文档）
    filtered_data = []
    for item in data:
        category = item.get('category', '')
        if category in EXCLUDE_CATEGORIES:
            continue
        filtered_data.append(item)
    print(f"过滤掉的需求数: {len(data) - len(filtered_data)}")
    data = filtered_data
    
    results = []
    t0 = time.time()
    
    # 准备所有匹配任务
    match_tasks = []
    for item in data:
        # 兼容两种格式：source_text在item层或在requirements中
        source = item.get('source_text', '')
        requirements = item.get('requirements', [])
        if not requirements:
            # 新格式：item本身就是一个需求
            req_text = item.get('requirement', '')
            category = item.get('category', '')
            if req_text:
                for p in products:
                    match_tasks.append((req_text, category, source, p))
        else:
            # 旧格式：requirements是列表
            for req in requirements:
                req_text = req.get('requirement', '')
                category = req.get('category', '')
                for p in products:
                    match_tasks.append((req_text, category, source, p))
    
    print(f"\n总任务数: {len(match_tasks)}, 开始并发处理...")
    
    # 并发执行所有匹配
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        match_results = list(executor.map(match_single_product, match_tasks))
    
    # 按需求整理结果（兼容新旧格式）
    task_idx = 0
    for item in data:
        source = item.get('source_text', '')
        requirements = item.get('requirements', [])
        
        if not requirements:
            # 新格式：item本身就是一个需求
            req_text = item.get('requirement', '')
            category = item.get('category', '')
            if not req_text:
                continue
                
            print(f"\n需求: {req_text[:30]}...")
            
            req_result = {
                "requirement": req_text,
                "category": category,
                "source_text": source,
                "matched_functions": []
            }
            
            # 收集该需求的所有产品匹配结果
            best_match = None
            best_score = -1
            best_product = None
            best_caps = None
            
            for p in products:
                result = match_results[task_idx]
                task_idx += 1
                
                if not result:
                    continue
                
                mf = result["mf"]
                score = result["score"]
                caps = result["caps"]
                product = result["product"]
                
                print(f"  {product}: {'是' if mf.get('is_product_function_matched') == '是' else '否'} (得分: {score})")
                
                if score > best_score:
                    best_score = score
                    best_match = mf
                    best_product = product
                    best_caps = caps
            
            # 处理最佳匹配结果
            if best_match:
                if best_score >= THRESHOLD:
                    best_match['is_product_function_matched'] = "是"
                    best_match['delivery_type'] = "标品功能"
                    
                    func_level = best_match.get('product_function_level', '')
                    if best_caps and func_level:
                        for cap in best_caps:
                            if func_level in cap.get('path', ''):
                                best_match['product_detail_source_text'] = cap.get('desc', '')
                                break
                    
                    req_result['matched_functions'].append(best_match)
                    print(f"  -> 已满足({best_product})，匹配度: {best_score}")
                else:
                    empty_match = {
                        "product_name": best_product,
                        "product_function_level": "",
                        "product_detail_source_text": "",
                        "founction_match_level": "0",
                        "is_product_function_matched": "否"
                    }
                    req_result['matched_functions'].append(empty_match)
                    print(f"  -> 不满足，风险评估...")
            
            results.append(req_result)
        else:
            # 旧格式：requirements是列表
            for req in requirements:
                req_text = req.get('requirement', '')
                category = req.get('category', '')
                
                print(f"\n需求: {req_text[:30]}...")
                
                req_result = {
                    "requirement": req_text,
                    "category": category,
                    "matched_functions": []
                }
                
                # 收集该需求的所有产品匹配结果
                best_match = None
                best_score = -1
                best_product = None
                best_caps = None
                
                for p in products:
                    result = match_results[task_idx]
                    task_idx += 1
                    
                    if not result:
                        continue
                    
                    mf = result["mf"]
                    score = result["score"]
                    caps = result["caps"]
                    product = result["product"]
                    
                    print(f"  {product}: {'是' if mf.get('is_product_function_matched') == '是' else '否'} (得分: {score})")
                    
                    if score > best_score:
                        best_score = score
                        best_match = mf
                        best_product = product
                        best_caps = caps
                
                # 处理最佳匹配结果
                if best_match:
                    if best_score >= THRESHOLD:
                        best_match['is_product_function_matched'] = "是"
                        best_match['delivery_type'] = "标品功能"
                        
                        func_level = best_match.get('product_function_level', '')
                        if best_caps and func_level:
                            for cap in best_caps:
                                if func_level in cap.get('path', ''):
                                    best_match['product_detail_source_text'] = cap.get('desc', '')
                                    break
                        
                        req_result['matched_functions'].append(best_match)
                        print(f"  -> 已满足({best_product})，匹配度: {best_score}")
                    else:
                        empty_match = {
                            "product_name": best_product,
                            "product_function_level": "",
                            "product_detail_source_text": "",
                            "founction_match_level": "0",
                            "is_product_function_matched": "否"
                        }
                        req_result['matched_functions'].append(empty_match)
                        print(f"  -> 不满足，风险评估...")
                
                results.append(req_result)
    
    # 并发执行风险评估
    t1 = time.time()
    print(f"\n=== 匹配阶段耗时: {t1-t0:.2f}秒 ===")
    print("\n=== 风险评估（并发）===")
    risk_tasks = []
    for req_result in results:
        mf = req_result.get('matched_functions', [])
        if mf and mf[0].get('is_product_function_matched') == '否':
            risk_tasks.append((req_result['requirement'], req_result['category'], mf[0]))
    
    if risk_tasks:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            risk_results = list(executor.map(risk_single_req, risk_tasks))
        
        task_idx = 0
        for req_result in results:
            mf = req_result.get('matched_functions', [])
            if mf and mf[0].get('is_product_function_matched') == '否':
                if task_idx < len(risk_results) and risk_results[task_idx]:
                    mf[0].update(risk_results[task_idx])
                task_idx += 1
    
    t2 = time.time()
    print(f"匹配耗时: {t1-t0:.2f}秒")
    print(f"风险评估耗时: {t2-t1:.2f}秒")
    print(f"总耗时: {t2-t0:.2f}秒")
    
    # 保存 JSON
    output_results = []
    for i, req_result in enumerate(results):
        idx = 0
        for item in data:
            for req in item.get('requirements', []):
                if idx == i:
                    req_result['source_text'] = item.get('source_text', '')
                    break
                idx += 1
        output_results.append(req_result)
    
    with open(OUT_PATH, 'w') as f:
        json.dump({"results": output_results}, f, ensure_ascii=False, indent=2)
    print(f"\n保存到: {OUT_PATH}")
    
    # 生成 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "匹配结果"
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = ["来源原文", "需求描述", "需求分类", "产品名称", "产品功能层级", "功能原文", "功能匹配度", "是否匹配", "交付方式", "需求质量度", "定制化工作说明", "是否开放需求", "风险应对策略"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # 写入数据
    row = 2
    for req_result in results:
        source_text = req_result.get('source_text', '')  # 来源原文
        requirement = req_result.get('requirement', '')  # 需求描述
        category = req_result.get('category', '')  # 需求分类
        
        for mf in req_result.get('matched_functions', []):
            ws.cell(row, 1, source_text).border = thin_border  # 来源原文
            ws.cell(row, 2, requirement).border = thin_border  # 需求描述
            ws.cell(row, 3, category).border = thin_border  # 需求分类
            ws.cell(row, 4, mf.get('product_name', '')).border = thin_border  # 产品名称
            ws.cell(row, 5, mf.get('product_function_level', '')).border = thin_border  # 产品功能层级
            ws.cell(row, 6, mf.get('product_detail_source_text', '') if mf.get('product_detail_source_text') else '').border = thin_border  # 功能原文
            ws.cell(row, 7, mf.get('founction_match_level', '')).border = thin_border  # 功能匹配度
            ws.cell(row, 8, mf.get('is_product_function_matched', '')).border = thin_border  # 是否匹配
            ws.cell(row, 9, mf.get('delivery_type', '')).border = thin_border  # 交付方式
            ws.cell(row, 10, mf.get('reqirement_quality_level', '')).border = thin_border  # 需求质量度
            ws.cell(row, 11, mf.get('customized_work_details', '') if mf.get('customized_work_details') else '').border = thin_border  # 定制化工作说明
            ws.cell(row, 12, mf.get('is_open_requirement', '') if mf.get('is_open_requirement') else '').border = thin_border  # 是否开放需求
            ws.cell(row, 13, mf.get('risk_management_strategy', '') if mf.get('risk_management_strategy') else '').border = thin_border  # 风险应对策略
            row += 1
    
    # 合并单元格逻辑：从左到右依次判断，只有当前面所有列都相同且当前列也相同时才合并
    for col in range(1, 14):  # 13列
        merge_start = None
        prev_values = []
        
        for r in range(2, row):  # 从第2行开始
            current_value = ws.cell(r, col).value or ""
            
            if col == 1:
                # 第1列：只需要比较当前列
                if merge_start is None:
                    merge_start = r
                    prev_value = current_value
                elif current_value == prev_value:
                    continue  # 继续合并
                else:
                    # 不相同，结束之前的合并
                    if r - merge_start > 1:  # 至少2行才合并
                        ws.merge_cells(start_row=merge_start, start_column=col, end_row=r-1, end_column=col)
                    merge_start = r
                    prev_value = current_value
            else:
                # 第2-13列：需要比较当前列和前面所有列
                current_row_values = [ws.cell(r, c).value or "" for c in range(1, col+1)]
                prev_row_values = [ws.cell(r-1, c).value or "" for c in range(1, col+1)]
                
                if merge_start is None:
                    merge_start = r
                    prev_values = current_row_values
                elif current_row_values == prev_row_values:
                    continue  # 继续合并
                else:
                    # 不相同，结束之前的合并
                    if r - merge_start > 1:
                        ws.merge_cells(start_row=merge_start, start_column=col, end_row=r-1, end_column=col)
                    merge_start = r
                    prev_values = current_row_values
        
        # 处理最后一组合并
        if merge_start and row - merge_start > 1:
            ws.merge_cells(start_row=merge_start, start_column=col, end_row=row-1, end_column=col)

    # 调整列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 40
    
    wb.save(OUT_EXCEL_PATH)
    print(f"Excel已生成: {OUT_EXCEL_PATH}")


    wb.save(OUT_EXCEL_PATH)
    print(f"Excel已生成: {OUT_EXCEL_PATH}")

if __name__ == "__main__":
    main()